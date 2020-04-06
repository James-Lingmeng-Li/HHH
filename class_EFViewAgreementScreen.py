import openpyxl, copy, re, math

from PyQt5 import QtCore, QtGui, QtWidgets

import HHHconf, HHHfunc, class_EFAmendPaymentScreen

class EFViewAgreementScreen(HHHconf.HHHWindowWidget):

    signal_agreementUpdated = QtCore.pyqtSignal(object)

    def __init__(self, selectedAgreement, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.selectedAgreement = str(selectedAgreement)
        self.addWidgets()
        self.getAgreementDetails()        
        self.getPayoutSchedule()                

    def showEvent(self, event):
        event.accept()
        try:
            self.childWindow.setFocus()
        except:
            pass

    def moveEvent(self, event):
        try:
            self.childWindow.move(self.geometry().x(), self.geometry().y())
        except:
            pass

    def closeEvent(self, event):
        try:
            self.childWindow.close()
        except:
            pass

    def addWidgets(self):
        self.setWindowTitle(self.title)
        self.setObjectName(HHHconf.windowObject)

        hBox_topButtons = QtWidgets.QHBoxLayout()
        hBox_topButtons.setContentsMargins(0, 0, 0, 0)
        hBox_topButtons.setSpacing(5)
        hBox_topButtons.setAlignment(QtCore.Qt.AlignTop)   
        self.vBox_contents.addLayout(hBox_topButtons)

        self.heading_EFViewClientNumber = QtWidgets.QLabel('Agreement Number: ' + self.selectedAgreement)
        self.heading_EFViewClientNumber.setStyleSheet(HHHconf.design_textMedium)
        hBox_topButtons.addWidget(self.heading_EFViewClientNumber)

        self.button_maintain_Agreement = QtWidgets.QPushButton('Maintain', clicked=self.maintain_Agreement, default=True)
        self.button_maintain_Agreement.setStyleSheet(HHHconf.design_smallButtonTransparent)
        self.button_maintain_Agreement.setFixedHeight(20)
        hBox_topButtons.addWidget(self.button_maintain_Agreement) 

        # Client Details
        self.clientDetailsBox = QtWidgets.QGroupBox()
        self.clientDetailsBox.setContentsMargins(10, 10, 10, 10)
        self.clientDetailsBox.setStyleSheet(HHHconf.design_GroupBox)
        self.vBox_contents.addWidget(self.clientDetailsBox)

        hBox_clientDetails = QtWidgets.QHBoxLayout()
        hBox_clientDetails.setContentsMargins(5, 5, 5, 5)
        hBox_clientDetails.setSpacing(10)
        hBox_clientDetails.setAlignment(QtCore.Qt.AlignTop)   
        self.clientDetailsBox.setLayout(hBox_clientDetails)

        vBox_clientDetails1 = QtWidgets.QVBoxLayout()
        vBox_clientDetails1.setContentsMargins(5, 5, 5, 5)
        vBox_clientDetails1.setSpacing(2)
        vBox_clientDetails1.setAlignment(QtCore.Qt.AlignTop)  
        hBox_clientDetails.addLayout(vBox_clientDetails1)
        vBox_clientDetails2 = QtWidgets.QVBoxLayout()
        vBox_clientDetails2.setContentsMargins(5, 5, 5, 5)
        vBox_clientDetails2.setSpacing(2)
        vBox_clientDetails2.setAlignment(QtCore.Qt.AlignTop)  
        hBox_clientDetails.addLayout(vBox_clientDetails2)   

        clientDetailsWidgetList1 = [ #[object_name, heading text, readOnly]
            [HHHconf.client_name, 'Client Name', True],  
            [HHHconf.street_address, 'Street Address', True], 
            [HHHconf.suburb, 'Suburb', True], 
            [HHHconf.state, 'State', True], 
            [HHHconf.postcode, 'Postcode', True]
        ]
        for detail in clientDetailsWidgetList1:
            headingName = 'Heading_EFView_' + detail[0]
            textboxName = 'Textbox_EFView_' + detail[0]
            self.headingName = QtWidgets.QLabel(detail[1], objectName=headingName)
            self.headingName.setStyleSheet(HHHconf.design_textSmall)
            vBox_clientDetails1.addWidget(self.headingName)
            self.textboxName = QtWidgets.QLineEdit('', objectName=textboxName, readOnly=detail[2])
            self.textboxName.resize(200, 20)
            self.textboxName.setStyleSheet(HHHconf.design_editBoxTwo)
            vBox_clientDetails1.addWidget(self.textboxName)

        clientDetailsWidgetList2 = [ #[object_name, heading text, readOnly]
            [HHHconf.abn, 'ABN', True], 
            [HHHconf.acn, 'ACN', True], 
            [HHHconf.contact_phone, 'Phone Number', True], 
            [HHHconf.contact_email, 'Email Address', True], 
            [HHHconf.contact_name, 'Contact Name', True]
        ]
        for detail in clientDetailsWidgetList2:
            headingName = 'Heading_EFView_' + detail[0]
            textboxName = 'Textbox_EFView_' + detail[0]
            self.headingName = QtWidgets.QLabel(detail[1], objectName=headingName)
            self.headingName.setStyleSheet(HHHconf.design_textSmall)
            vBox_clientDetails2.addWidget(self.headingName)
            self.textboxName = QtWidgets.QLineEdit('', objectName=textboxName, readOnly=detail[2])
            self.textboxName.resize(200, 20)
            self.textboxName.setStyleSheet(HHHconf.design_editBoxTwo)
            vBox_clientDetails2.addWidget(self.textboxName)

        # Agreement Details
        hBox_agreementDetails = QtWidgets.QHBoxLayout()
        hBox_agreementDetails.setContentsMargins(0, 0, 0, 0)
        hBox_agreementDetails.setSpacing(5)
        hBox_agreementDetails.setAlignment(QtCore.Qt.AlignTop)   
        self.vBox_contents.addLayout(hBox_agreementDetails)       

        self.agreementDetailsBox = QtWidgets.QGroupBox()
        self.agreementDetailsBox.setFixedWidth(250)
        self.agreementDetailsBox.setMinimumHeight(300)
        self.agreementDetailsBox.setContentsMargins(10, 10, 10, 10)
        self.agreementDetailsBox.setStyleSheet(HHHconf.design_GroupBox)
        hBox_agreementDetails.addWidget(self.agreementDetailsBox)

        vBox_agreementDetails = QtWidgets.QVBoxLayout()
        vBox_agreementDetails.setContentsMargins(10, 10, 0, 10)
        vBox_agreementDetails.setSpacing(0)
        vBox_agreementDetails.setAlignment(QtCore.Qt.AlignTop)   
        self.agreementDetailsBox.setLayout(vBox_agreementDetails)

        self.list_agreementDetails = QtWidgets.QListWidget()
        self.list_agreementDetails.setFocusPolicy(QtCore.Qt.NoFocus)
        self.list_agreementDetails.setStyleSheet(HHHconf.design_listTwo)
        self.list_agreementDetails.setSelectionMode(QtWidgets.QAbstractItemView.NoSelection)
        self.list_agreementDetails.setVerticalScrollMode(QtWidgets.QAbstractItemView.ScrollPerPixel)
        self.list_agreementDetails.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOn)
        self.list_agreementDetails.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        vBox_agreementDetails.addWidget(self.list_agreementDetails)

        agreementDetailsWidgetList = [ #[object_name, heading text, readOnly]
            [HHHconf.original_balance, 'Original Balance', True], 
            [HHHconf.balloon_amount, 'Balloon Amount', True], 
            [HHHconf.periodic_fee, 'Periodic Fee', True], 
            [HHHconf.interest_rate, 'Interest Rate (p.a.)', True], 
            [HHHconf.periodic_repayment, 'Periodic Repayment', True], 
            [HHHconf.periods_per_year, 'Periods Per Year', True], 
            [HHHconf.total_periods, 'Total Periods', True], 
            [HHHconf.settlement_date, 'Settlement Date (' + HHHconf.dateFormat + ')', True], 
            [HHHconf.agreement_start_date, 'Agreement Start Date (' + HHHconf.dateFormat + ')', True], 
            [HHHconf.mtx_number, 'Mtx Buyer Number', True], 
            [HHHconf.bsb, 'Bank BSB', True], 
            [HHHconf.acc, 'Bank ACC', True]
        ]
        for detail in agreementDetailsWidgetList:
            containingWidget = QtWidgets.QWidget()  
            item_detail = QtWidgets.QListWidgetItem()
            item_detail.setData(QtCore.Qt.UserRole, 'agreementDetail_' + detail[0])
            item_detail.setTextAlignment(QtCore.Qt.AlignCenter)

            vBox_agreementDetailsList = QtWidgets.QVBoxLayout()
            vBox_agreementDetailsList.setContentsMargins(0, 0, 0, 6)
            vBox_agreementDetailsList.setSpacing(0)
            vBox_agreementDetailsList.setAlignment(QtCore.Qt.AlignTop)   
            containingWidget.setLayout(vBox_agreementDetailsList)

            headingName = 'Heading_EFView_' + detail[0]
            textboxName = 'Textbox_EFView_' + detail[0]
            heading_detail = QtWidgets.QLabel(detail[1], objectName=headingName)
            heading_detail.setStyleSheet(HHHconf.design_textSmall)
            vBox_agreementDetailsList.addWidget(heading_detail)
            edit_detail = QtWidgets.QLineEdit('', objectName=textboxName, readOnly=detail[2])
            edit_detail.resize(200, 20)
            edit_detail.setStyleSheet(HHHconf.design_editBoxTwo)
            vBox_agreementDetailsList.addWidget(edit_detail)

            item_detail.setFlags(item_detail.flags() & ~QtCore.Qt.ItemIsSelectable)
            self.list_agreementDetails.addItem(item_detail)
            self.list_agreementDetails.setItemWidget(item_detail, containingWidget)
            item_detail.setSizeHint(QtCore.QSize(containingWidget.width(), containingWidget.height()))

        # Agreement Notes
        vBox_agreementNotes = QtWidgets.QVBoxLayout()
        vBox_agreementNotes.setContentsMargins(0, 0, 0, 0)
        vBox_agreementNotes.setSpacing(5)
        vBox_agreementNotes.setAlignment(QtCore.Qt.AlignTop)   
        hBox_agreementDetails.addLayout(vBox_agreementNotes)

        self.agreementNotesList = QtWidgets.QListWidget()
        self.agreementNotesList.setFocusPolicy(QtCore.Qt.NoFocus)
        self.agreementNotesList.setStyleSheet(HHHconf.design_listTwo)
        self.agreementNotesList.setVerticalScrollMode(QtWidgets.QAbstractItemView.ScrollPerPixel)
        vBox_agreementNotes.addWidget(self.agreementNotesList)
        for key in [HHHconf.agreement_notes_asset, HHHconf.agreement_notes_account, HHHconf.agreement_notes_misc]:
            noteName = 'Heading_EFView_' + str(key)
            textEditName = 'TextEdit_EFView_' +  str(key)
            noteHeading = HHHconf.dict_toSQL_HHH_EF_Agreement_Info[key]
            item = QtWidgets.QListWidgetItem()
            itemWidget = QtWidgets.QWidget()
            itemLayout = QtWidgets.QVBoxLayout()
            itemLayout.setContentsMargins(0, 0, 5, 0)
            itemLayout.setSpacing(6)
            itemWidget.setLayout(itemLayout)
            noteCategory = QtWidgets.QLabel(re.sub(r'(?<=\w)([A-Z])', r' \1', noteHeading), objectName=noteName)
            noteCategory.setStyleSheet(HHHconf.design_textSmall)
            itemLayout.addWidget(noteCategory)
            noteText = HHHconf.HHHTextEditWidget(self, objName=textEditName, placeholderText=HHHconf.widgetPlaceHolderText_notes)
            noteText.setMaximumHeight(113)
            itemLayout.addWidget(noteText)
            item.setSizeHint(itemWidget.sizeHint())
            self.agreementNotesList.addItem(item)
            self.agreementNotesList.setItemWidget(item, itemWidget)

       # Payout Schedule Table
        self.EFViewPayoutScheduleTable = HHHconf.HHHTableWidget(cornerWidget='copy', contextMenu=self.EFViewPayoutScheduleTableMenu, doubleClicked=self.EFViewPayoutScheduleTable_doubleClicked)
        self.vBox_contents.addWidget(self.EFViewPayoutScheduleTable)

        self.button_update.clicked.connect(self.updateAgreement)
    
    def getAgreementDetails(self):
        # Get Agreement Data from db
        self.df_EFAgreementDetails = HHHfunc.equipmentFinanceEngine.getAgreementDetails([self.selectedAgreement])
        self.clientNumber = self.df_EFAgreementDetails.at[0, HHHconf.client_number]
        
        # Get Client Data from db
        self.df_clientDetails = HHHfunc.equipmentFinanceEngine.getClientDetails([self.clientNumber])

        # Display Client data
        for key in self.df_clientDetails.keys():
            if self.findChild(QtWidgets.QLineEdit, 'Textbox_EFView_' + key) is None:
                continue
            else:
                self.findChild(QtWidgets.QLineEdit, 'Textbox_EFView_' + key).setText(str(self.df_clientDetails.iloc[0][key]))
                self.findChild(QtWidgets.QLineEdit, 'Textbox_EFView_' + key).setCursorPosition(0)

        # Display Agreement data
        for key in self.df_EFAgreementDetails.keys():
            if self.findChild(QtWidgets.QLineEdit, 'Textbox_EFView_' + key) is None:
                continue
            if key == HHHconf.original_balance or key == HHHconf.balloon_amount or key == HHHconf.periodic_repayment or key == HHHconf.periodic_fee: # round up to nearest cent
                 self.findChild(QtWidgets.QLineEdit, 'Textbox_EFView_' + key).setText(HHHconf.moneyFormat.format(math.ceil(self.df_EFAgreementDetails.iloc[0][key] * 100) / 100))
            elif key == HHHconf.interest_rate:
                self.findChild(QtWidgets.QLineEdit, 'Textbox_EFView_' + key).setText('{:0,.2f}%'.format(self.df_EFAgreementDetails.iloc[0][key] * 100))
            elif key == HHHconf.agreement_start_date or key == HHHconf.settlement_date:
                self.findChild(QtWidgets.QLineEdit, 'Textbox_EFView_' + key).setText(self.df_EFAgreementDetails.iloc[0][key].strftime(HHHconf.dateFormat2))
            elif key == HHHconf.total_periods:
                self.findChild(QtWidgets.QLineEdit, 'Textbox_EFView_' + key).setText('{:0.0f}'.format(self.df_EFAgreementDetails.iloc[0][key]))
            else:
                self.findChild(QtWidgets.QLineEdit, 'Textbox_EFView_' + key).setText(str(self.df_EFAgreementDetails.iloc[0][key]))
        for key in [HHHconf.agreement_notes_asset, HHHconf.agreement_notes_account, HHHconf.agreement_notes_misc]:
            self.findChild(QtWidgets.QTextEdit, 'TextEdit_EFView_' + key).setPlainText(self.df_EFAgreementDetails.iloc[0][key])

    def getPayoutSchedule(self):
        self.df_paymentDetails = HHHfunc.equipmentFinanceEngine.getPaymentDetails([self.selectedAgreement])
        df_paymentDetails = copy.deepcopy(self.df_paymentDetails[[HHHconf.payment_number, HHHconf.repayment_date, HHHconf.opening_balance, HHHconf.interest_component, HHHconf.principal_component, HHHconf.periodic_repayment, HHHconf.closing_balance]]) # display-only

        df_paymentDetails.insert(loc=1, column='OutstandingSubordinates', value=0)        
        # Calculate outstanding subordinates for each payment (data does not exist in db)
        df_currentSubordinates = HHHfunc.equipmentFinanceEngine.getCurrentSubordinates(selectedAgreement=self.selectedAgreement, collapseGST=True)
        df_outstandingSubordinates = df_currentSubordinates[df_currentSubordinates[HHHconf.subordinate_status].isin(HHHconf.outstandingStatuses)]
        # Add outstanding subordinates to df
        df_paymentDetails['OutstandingSubordinates'] = df_paymentDetails[HHHconf.payment_number].map(lambda x: df_outstandingSubordinates[df_outstandingSubordinates[HHHconf.payment_number] == x][HHHconf.subordinate_amount].sum())

        # Format columns
        for col in df_paymentDetails:
            if col == HHHconf.payment_number:
                continue
            elif col == HHHconf.repayment_date:
                df_paymentDetails[col] = df_paymentDetails[col].dt.strftime(HHHconf.dateFormat2)
            elif col == HHHconf.periodic_repayment:
                df_paymentDetails[col] = df_paymentDetails[col].map(lambda x: HHHconf.moneyFormat.format(math.ceil(abs(x) * 100) / 100))
            else:
                df_paymentDetails[col] = df_paymentDetails[col].map(HHHconf.moneyFormat.format)

        df_paymentDetails.rename({HHHconf.payment_number:'Payment Number', 'OutstandingSubordinates':'Outstanding', HHHconf.repayment_date:'Date', HHHconf.opening_balance:'Opening Balance', HHHconf.interest_component:'Interest', HHHconf.principal_component:'Principal', HHHconf.periodic_repayment:'Repayment', HHHconf.closing_balance:'Closing Balance'}, axis='columns', inplace=True)
        self.EFViewPayoutScheduleTable.setModel(df_paymentDetails)

    def extract_Agreement(self):
        self.dict_EFAgreementDetails = self.df_EFAgreementDetails.to_dict('records')[0]
        self.dict_clientDetails = self.df_clientDetails.to_dict('records')[0]
        print(self.dict_EFAgreementDetails)
        print(self.dict_clientDetails)

        self.df_transactionDetails = HHHfunc.equipmentFinanceEngine.getCurrentSubordinates(selectedAgreement=self.selectedAgreement)

        print(self.df_transactionDetails)

        wb_agreementStatement = openpyxl.Workbook()

        # Set up cell styles
        for style, styleValues in HHHconf.dict_statementStyles.items():
            applyStyle = openpyxl.styles.NamedStyle(name=style)
            applyStyle.font = openpyxl.styles.Font(name=styleValues['font'], bold=styleValues.get('bold'), size=styleValues['size'], color=styleValues['color'])
            applyStyle.alignment = openpyxl.styles.Alignment(horizontal=styleValues['horizontal-alignment'], vertical=styleValues['vertical-alignment'])
            if styleValues.get('background-color') is not None:
                applyStyle.fill = openpyxl.styles.PatternFill(fill_type='solid', start_color=styleValues['background-color'], end_color=styleValues['background-color'])
            wb_agreementStatement.add_named_style(applyStyle)

        ws_agreementStatement = wb_agreementStatement.active
        statementName = ws_agreementStatement.cell(row=1, column=1, value='Moneytech Equipment Finance Statement')
        statementName.style = HHHconf.styleHeading
        ws_agreementStatement.merge_cells('A1:D1')
        statementDate = ws_agreementStatement.cell(row=1, column=5, value='=TODAY()')
        statementDate.style = HHHconf.styleHeading
        statementDate.number_format = 'dd-mmm-yyyy'

        dict_statementClientDetails = {
            'A3':{'displayText':'YOUR DETAILS', 'displayStyle':HHHconf.styleHeader, 'displayMerge':'E3'}, 
            'A4':{'displayText':'Name', 'displayStyle':HHHconf.styleLabel, 'offset':{'row':0, 'column':1}, 'value':self.dict_clientDetails[HHHconf.client_name], 'valueStyle':HHHconf.styleValue}, 
            'A5':{'displayText':'ABN', 'displayStyle':HHHconf.styleLabel, 'offset':{'row':0, 'column':1}, 'value':self.dict_clientDetails[HHHconf.abn], 'valueStyle':HHHconf.styleValue}, 
            'A6':{'displayText':'ACN', 'displayStyle':HHHconf.styleLabel, 'offset':{'row':0, 'column':1}, 'value':self.dict_clientDetails[HHHconf.acn], 'valueStyle':HHHconf.styleValue}, 
            'A7':{'displayText':'Phone', 'displayStyle':HHHconf.styleLabel, 'offset':{'row':0, 'column':1}, 'value':self.dict_clientDetails[HHHconf.contact_phone], 'valueStyle':HHHconf.styleValue}, 
            'A9':{'displayText':'Contact Name', 'displayStyle':HHHconf.styleLabel, 'offset':{'row':0, 'column':1}, 'value':self.dict_clientDetails[HHHconf.contact_name], 'valueStyle':HHHconf.styleValue}, 
            'D4':{'displayText':'Street Address', 'displayStyle':HHHconf.styleLabel, 'offset':{'row':0, 'column':1}, 'value':self.dict_clientDetails[HHHconf.street_address], 'valueStyle':HHHconf.styleValue}, 
            'D5':{'displayText':'Suburb', 'displayStyle':HHHconf.styleLabel, 'offset':{'row':0, 'column':1}, 'value':self.dict_clientDetails[HHHconf.suburb], 'valueStyle':HHHconf.styleValue}, 
            'D6':{'displayText':'State', 'displayStyle':HHHconf.styleLabel, 'offset':{'row':0, 'column':1}, 'value':self.dict_clientDetails[HHHconf.state], 'valueStyle':HHHconf.styleValue}, 
            'D7':{'displayText':'Postcode', 'displayStyle':HHHconf.styleLabel, 'offset':{'row':0, 'column':1}, 'value':self.dict_clientDetails[HHHconf.postcode], 'valueStyle':HHHconf.styleValue}, 
            'D9':{'displayText':'Email', 'displayStyle':HHHconf.styleLabel, 'offset':{'row':0, 'column':1}, 'value':self.dict_clientDetails[HHHconf.contact_email], 'valueStyle':HHHconf.styleValue}
        }
        for cell, dict_value in dict_statementClientDetails.items():
            currentDisplayCell = ws_agreementStatement[cell]
            currentDisplayCell.value = dict_value['displayText']
            currentDisplayCell.style = dict_value['displayStyle']
            if dict_value.get('displayMerge') is not None:
                ws_agreementStatement.merge_cells(cell + ':' + dict_value['displayMerge'])

            if dict_value.get('value') is not None:
                currentValueCell = ws_agreementStatement[cell].offset(row=dict_value['offset']['row'], column=dict_value['offset']['column'])
                currentValueCell.value = dict_value['value']
                currentValueCell.style = dict_value['valueStyle']
                if dict_value.get('valueFormat') is not None:
                    currentValueCell.number_format = dict_value['valueFormat']

        dict_statementAgreementDetails = {
            'A11':{'displayText':'AGREEMENT DETAILS', 'displayStyle':HHHconf.styleHeader, 'displayMerge':'E11'}, 
            'A12':{'displayText':'Agreement Number:', 'displayStyle':HHHconf.styleLabel, 'offset':{'row':0, 'column':1}, 'value':self.dict_EFAgreementDetails[HHHconf.agreement_number], 'valueStyle':HHHconf.styleValue}, 
            'A14':{'displayText':'Original Balance:', 'displayStyle':HHHconf.styleLabel, 'offset':{'row':0, 'column':1}, 'value':self.dict_EFAgreementDetails[HHHconf.original_balance], 'valueStyle':HHHconf.styleValue, 'valueFormat':'$#, ##0.00;-$#, ##0.00'}, 
            'A15':{'displayText':'Periodic Repayment:', 'displayStyle':HHHconf.styleLabel, 'offset':{'row':0, 'column':1}, 'value':self.dict_EFAgreementDetails[HHHconf.periodic_repayment], 'valueStyle':HHHconf.styleValue, 'valueFormat':'$#, ##0.00;-$#, ##0.00'}, 
            'A16':{'displayText':'Repayments Per Year:', 'displayStyle':HHHconf.styleLabel, 'offset':{'row':0, 'column':1}, 'value':self.dict_EFAgreementDetails[HHHconf.periods_per_year], 'valueStyle':HHHconf.styleValue}, 
            'A17':{'displayText':'Total Repayments:', 'displayStyle':HHHconf.styleLabel, 'offset':{'row':0, 'column':1}, 'value':self.dict_EFAgreementDetails[HHHconf.total_periods], 'valueStyle':HHHconf.styleValue}, 
            'D12':{'displayText':'Mtx (internal use):', 'displayStyle':HHHconf.styleLabel, 'offset':{'row':0, 'column':1}, 'value':self.dict_EFAgreementDetails[HHHconf.mtx_number], 'valueStyle':HHHconf.styleValue}, 
            'D14':{'displayText':'Settlement Date:', 'displayStyle':HHHconf.styleLabel, 'offset':{'row':0, 'column':1}, 'value':self.dict_EFAgreementDetails[HHHconf.settlement_date], 'valueStyle':HHHconf.styleValue, 'valueFormat':'dd-mmm-yyyy'}, 
            'D15':{'displayText':'First Payment Date:', 'displayStyle':HHHconf.styleLabel, 'offset':{'row':0, 'column':1}, 'value':self.dict_EFAgreementDetails[HHHconf.agreement_start_date], 'valueStyle':HHHconf.styleValue, 'valueFormat':'dd-mmm-yyyy'}, 
            'D16':{'displayText':'Subordinate Charges:', 'displayStyle':HHHconf.styleLabel, 'offset':{'row':0, 'column':1}, 'value':'Add total charges incurred', 'valueStyle':HHHconf.styleValue, 'valueFormat':'$#, ##0.00;-$#, ##0.00'}, 
            'D17':{'displayText':'Charges Outstanding:', 'displayStyle':HHHconf.styleLabel, 'offset':{'row':0, 'column':1}, 'value':'Add total unpaid charges', 'valueStyle':HHHconf.styleValue, 'valueFormat':'$#, ##0.00;-$#, ##0.00'}
        }
        for cell, dict_value in dict_statementAgreementDetails.items():
            currentDisplayCell = ws_agreementStatement[cell]
            currentDisplayCell.value = dict_value['displayText']
            currentDisplayCell.style = dict_value['displayStyle']
            if dict_value.get('displayMerge') is not None:
                ws_agreementStatement.merge_cells(cell + ':' + dict_value['displayMerge'])

            if dict_value.get('value') is not None:
                currentValueCell = ws_agreementStatement[cell].offset(row=dict_value['offset']['row'], column=dict_value['offset']['column'])
                currentValueCell.value = dict_value['value']
                currentValueCell.style = dict_value['valueStyle']
                if dict_value.get('valueFormat') is not None:
                    currentValueCell.number_format = dict_value['valueFormat']

        wb_agreementStatement.save('C:\\Users\\Hugh.huang\\Desktop\\testStatement.xlsx')

    def maintain_Agreement(self):
        self.childWindow = maintainAgreementScreen(username=self.username, geometry=self.geometry(), padding={'left':20, 'top':100, 'right':20, 'bottom':100}, title='Maintain Agreement Screen', parent=self)
        self.childWindow.show()
        self.childWindow.activateWindow()
        self.childWindow.setFocus()

    def updateAgreement(self):
        cnxn = HHHfunc.mainEngine.establishSqlConnection()
        cursor = cnxn.cursor()

        # Update Notes
        dict_agreementNotes = {}
        for key in [HHHconf.agreement_notes_asset, HHHconf.agreement_notes_account, HHHconf.agreement_notes_misc]:
            currentNotes =  self.findChild(QtWidgets.QTextEdit, 'TextEdit_EFView_' + key).toPlainText()
            if len(currentNotes) == 0:
                currentNotes = None
            dict_agreementNotes[key] = currentNotes
        cursor.execute('UPDATE ' + HHHconf.db_agreementInfo_name + ' SET ' + HHHconf.dict_toSQL_HHH_EF_Agreement_Info[HHHconf.agreement_notes_asset] + ' = ?, ' + HHHconf.dict_toSQL_HHH_EF_Agreement_Info[HHHconf.agreement_notes_account] + ' = ?, ' + HHHconf.dict_toSQL_HHH_EF_Agreement_Info[HHHconf.agreement_notes_misc] + ' = ? WHERE ' + HHHconf.dict_toSQL_HHH_EF_Agreement_Info[HHHconf.agreement_number] + ' = ?', dict_agreementNotes[HHHconf.agreement_notes_asset], dict_agreementNotes[HHHconf.agreement_notes_account], dict_agreementNotes[HHHconf.agreement_notes_misc], self.selectedAgreement)
        cnxn.commit()
        self.signal_agreementUpdated.emit({HHHconf.events_EventTime: None, HHHconf.events_UserName: HHHconf.username_PC, HHHconf.events_Event: HHHconf.dict_eventCategories['equipmentFinance'], HHHconf.events_EventDescription: 'Updated Agreement Notes | ' + HHHconf.agreement_notes_asset + ': ' + (dict_agreementNotes[HHHconf.agreement_notes_asset] if dict_agreementNotes[HHHconf.agreement_notes_asset] else '') + ' | ' + HHHconf.agreement_notes_account + ': ' + (dict_agreementNotes[HHHconf.agreement_notes_account] if dict_agreementNotes[HHHconf.agreement_notes_account] else '') + ' | ' + HHHconf.agreement_notes_misc + ': ' + (dict_agreementNotes[HHHconf.agreement_notes_misc] if dict_agreementNotes[HHHconf.agreement_notes_misc] else '') + ' | '})
        return HHHconf.widgetEffects.flashMessage(self.heading_status, duration=1000, message='Updated Agreement Notes', messageType='success')

    def EFViewPayoutScheduleTableMenu(self, pos):
        # TODO - repurpose context menu or delete
        # selectedPayment = str(self.EFViewPayoutScheduleTable.model().index(self.EFViewPayoutScheduleTable.currentIndex().row(), 0).data(QtCore.Qt.DisplayRole))
        # menu = QtWidgets.QMenu()
        # menu.setStyleSheet(HHHconf.design_tableMenu)
        # Action_amendPayment = menu.addAction('Amend Payment')
        # action = menu.exec_(QtGui.QCursor.pos())
        # if action == Action_amendPayment:
        #     self.EFViewPayoutScheduleTable_doubleClicked()
        pass

    def EFViewPayoutScheduleTable_doubleClicked(self):
        selectedPayment = str(self.EFViewPayoutScheduleTable.model().index(self.EFViewPayoutScheduleTable.currentIndex().row(), 0).data(QtCore.Qt.DisplayRole))
        self.EFAmendPaymentScreen(selectedPayment) 

    def EFAmendPaymentScreen(self, selectedPayment):
        self.childWindow =  class_EFAmendPaymentScreen.EFAmendPaymentScreen(selectedAgreement=self.selectedAgreement, selectedPayment=selectedPayment, username=self.username, geometry=self.geometry(), padding={'left':0, 'top':20, 'right':0, 'bottom':220}, title='Amend Payment Details', updateButton='Update Payment Notes', parent=self)
        self.childWindow.show()
        self.childWindow.activateWindow()
        self.childWindow.setFocus()
        self.childWindow.signal_paymentUpdated.connect(self.getPayoutSchedule)
        self.childWindow.signal_paymentUpdated.connect(lambda x: self.signal_agreementUpdated.emit(x))

#incomplete
class maintainAgreementScreen(HHHconf.HHHWindowWidget):
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.dict_sections = {
            'statements':{
                'heading':'Statements', 
                'widget':self.statementBox
            }, 
            'payoutCalculations':{
                'heading':'Payout Calculations', 
                'widget':self.payoutBox
            }, 
            'clientUpdate':{
                'heading':'Client', 
                'widget':self.clientBox
            }, 
            'agreementUpdate':{
                'heading':'Agreement', 
                'widget':self.agreementBox
            }, 
            'termination':{
                'heading':'Terminate', 
                'widget':self.terminateBox
            }, 
        }
        self.addWidgets()        

    def addWidgets(self):
        hBox_topButtons = QtWidgets.QHBoxLayout()
        hBox_topButtons.setContentsMargins(0, 0, 0, 0)
        hBox_topButtons.setSpacing(5)
        hBox_topButtons.setAlignment(QtCore.Qt.AlignTop)   
        self.vBox_contents.addLayout(hBox_topButtons)

        self.heading_maintain = QtWidgets.QLabel('Maintain Agreement')
        self.heading_maintain.setStyleSheet(HHHconf.design_textMedium)
        hBox_topButtons.addWidget(self.heading_maintain)

        hBox_contents = QtWidgets.QHBoxLayout()
        hBox_contents.setContentsMargins(0, 0, 0, 0)
        hBox_contents.setSpacing(5)
        hBox_contents.setAlignment(QtCore.Qt.AlignTop)   
        self.vBox_contents.addLayout(hBox_contents) 

        # Select Section to jump to (list for ease of add more)
        self.selectSectionsList = QtWidgets.QListWidget()
        self.selectSectionsList.setFocusPolicy(QtCore.Qt.NoFocus)
        self.selectSectionsList.setStyleSheet(HHHconf.design_listTwo)
        self.selectSectionsList.setVerticalScrollMode(QtWidgets.QAbstractItemView.ScrollPerPixel)
        self.selectSectionsList.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOn)
        self.selectSectionsList.setFixedWidth(150)
        self.selectSectionsList.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        self.selectSectionsList.currentItemChanged.connect(self.goToSection)
        hBox_contents.addWidget(self.selectSectionsList)

        # Main contents list
        self.mainContentsList = QtWidgets.QListWidget()
        self.mainContentsList.setFocusPolicy(QtCore.Qt.NoFocus)
        self.mainContentsList.setStyleSheet(HHHconf.design_listTwo)
        self.mainContentsList.setSelectionMode(QtWidgets.QAbstractItemView.NoSelection)
        self.mainContentsList.setVerticalScrollMode(QtWidgets.QAbstractItemView.ScrollPerPixel)
        self.mainContentsList.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOn)
        self.mainContentsList.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        hBox_contents.addWidget(self.mainContentsList)

        for obj, options in self.dict_sections.items():
            item_section = QtWidgets.QListWidgetItem(options['heading'])
            item_section.setData(QtCore.Qt.UserRole, 'section_' + obj)
            item_section.setTextAlignment(QtCore.Qt.AlignCenter)
            item_section.setSizeHint(QtCore.QSize(self.selectSectionsList.width()-12, 50))
            self.selectSectionsList.addItem(item_section)

            boxWidget = options['widget']()
            item = QtWidgets.QListWidgetItem()
            item.setData(QtCore.Qt.UserRole, 'box_' + obj)
            item.setSizeHint(QtCore.QSize(self.mainContentsList.sizeHint().width(), boxWidget.height() + 10))
            item.setFlags(item.flags() & ~QtCore.Qt.ItemIsSelectable)
            self.mainContentsList.addItem(item)
            self.mainContentsList.setItemWidget(item, boxWidget)


    def goToSection(self, selectedSection, prevSection):
        print(selectedSection)
        print(prevSection)

    def statementBox(self):
        box_container = QtWidgets.QGroupBox()
        box_container.setContentsMargins(0, 0, 0, 0)
        box_container.setStyleSheet(HHHconf.design_GroupBox.replace(HHHconf.secondaryBackgroundColor, HHHconf.transparentWidgetColor))

        vBox_container = QtWidgets.QVBoxLayout()
        vBox_container.setContentsMargins(0, 0, 0, 0)
        vBox_container.setSpacing(5)
        vBox_container.setAlignment(QtCore.Qt.AlignTop)   
        box_container.setLayout(vBox_container)

        box_statements = QtWidgets.QGroupBox()
        box_statements.setContentsMargins(5, 5, 5, 5)
        box_statements.setStyleSheet(HHHconf.design_GroupBox)
        box_statements.setFixedHeight(100)
        vBox_container.addWidget(box_statements)

        vBox_contents = QtWidgets.QVBoxLayout()
        vBox_contents.setContentsMargins(5, 5, 5, 5)
        vBox_contents.setSpacing(5)
        vBox_contents.setAlignment(QtCore.Qt.AlignTop)   
        box_statements.setLayout(vBox_contents)

        heading = QtWidgets.QLabel(self.dict_sections['statements']['heading'])
        heading.setStyleSheet(HHHconf.design_textMedium)
        vBox_contents.addWidget(heading)
        return box_container

    def payoutBox(self):
        box_payout = QtWidgets.QGroupBox()
        box_payout.setContentsMargins(5, 5, 5, 5)
        box_payout.setStyleSheet(HHHconf.design_GroupBox)
        box_payout.setFixedHeight(200)

        vBox_container = QtWidgets.QVBoxLayout()
        vBox_container.setContentsMargins(5, 5, 5, 5)
        vBox_container.setSpacing(5)
        vBox_container.setAlignment(QtCore.Qt.AlignTop)   
        box_payout.setLayout(vBox_container)

        heading = QtWidgets.QLabel(self.dict_sections['payoutCalculations']['heading'])
        heading.setStyleSheet(HHHconf.design_textMedium)
        vBox_container.addWidget(heading)
        return box_payout
    
    def clientBox(self):
        box_client = QtWidgets.QGroupBox()
        box_client.setContentsMargins(5, 5, 5, 5)
        box_client.setStyleSheet(HHHconf.design_GroupBox)
        box_client.setFixedHeight(150)

        vBox_container = QtWidgets.QVBoxLayout()
        vBox_container.setContentsMargins(5, 5, 5, 5)
        vBox_container.setSpacing(5)
        vBox_container.setAlignment(QtCore.Qt.AlignTop)   
        box_client.setLayout(vBox_container)

        heading = QtWidgets.QLabel(self.dict_sections['clientUpdate']['heading'])
        heading.setStyleSheet(HHHconf.design_textMedium)
        vBox_container.addWidget(heading)
        return box_client
    
    def agreementBox(self):
        box_agreement = QtWidgets.QGroupBox()
        box_agreement.setContentsMargins(5, 5, 5, 5)
        box_agreement.setStyleSheet(HHHconf.design_GroupBox)
        box_agreement.setFixedHeight(100)

        vBox_container = QtWidgets.QVBoxLayout()
        vBox_container.setContentsMargins(5, 5, 5, 5)
        vBox_container.setSpacing(5)
        vBox_container.setAlignment(QtCore.Qt.AlignTop)   
        box_agreement.setLayout(vBox_container)

        heading = QtWidgets.QLabel(self.dict_sections['agreementUpdate']['heading'])
        heading.setStyleSheet(HHHconf.design_textMedium)
        vBox_container.addWidget(heading)
        return box_agreement

    def terminateBox(self):
        box_termination = QtWidgets.QGroupBox()
        box_termination.setContentsMargins(5, 5, 5, 5)
        box_termination.setStyleSheet(HHHconf.design_GroupBox)
        box_termination.setFixedHeight(100)

        vBox_container = QtWidgets.QVBoxLayout()
        vBox_container.setContentsMargins(5, 5, 5, 5)
        vBox_container.setSpacing(5)
        vBox_container.setAlignment(QtCore.Qt.AlignTop)   
        box_termination.setLayout(vBox_container)

        heading = QtWidgets.QLabel(self.dict_sections['termination']['heading'])
        heading.setStyleSheet(HHHconf.design_textMedium)
        vBox_container.addWidget(heading)
        return box_termination